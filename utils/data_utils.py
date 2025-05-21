# Import necessary libraries
import pandas as pd
import torch
from torch.utils.data import Dataset
from sklearn.preprocessing import LabelEncoder
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.metrics import auc

# List of categorical columns used as features in the model
categorical_cols = [
    "src_user", "src_domain", "dst_user", "dst_domain",
    "src_comp", "dst_comp", "auth_type", "logon_type",
    "orientation", "success"
]

def load_and_encode_data(csv_path, start=2500000, end=2580000):
    """
    Loads data from a CSV file, slices a specified range of rows,
    assigns column names, and encodes categorical columns using LabelEncoder.

    Args:
        csv_path (str): Path to the CSV file.
        start (int): Starting row index to read (inclusive).
        end (int): Ending row index to read (exclusive); use -1 to read until end.

    Returns:
        df (DataFrame): Processed and encoded dataframe.
        encoders (dict): Dictionary of fitted LabelEncoders for each categorical column.
    """
    # Read the CSV file as string to prevent unwanted type conversion
    df = pd.read_csv(csv_path, header=None, dtype=str, low_memory=False)

    # Handle case where end is -1 (use until end of DataFrame)
    if end == -1:
        end = df.shape[0]

    # Slice the DataFrame to the desired row range
    df = df[start:end].reset_index(drop=True)

    # Assign meaningful column names
    df.columns = [
        "time", "src_user", "src_domain", "dst_user", "dst_domain",
        "src_comp", "dst_comp", "auth_type", "logon_type", "orientation",
        "success", "is_redteam"
    ]

    encoders = {}

    # Encode each categorical column with LabelEncoder
    for col in categorical_cols:
        df[col] = df[col].astype(str).fillna("UNK")  # Replace NaNs with 'UNK'
        le = LabelEncoder()
        df[col] = le.fit_transform(df[col])
        encoders[col] = le  # Save encoder for future inverse transforms

    return df, encoders

class AuthSequenceDataset(Dataset):
    """
    Custom PyTorch Dataset for sequence prediction using authentication logs.

    Each sample consists of a sequence of categorical features and the target
    is the next row (i+seq_len) in the sequence.

    Args:
        df (DataFrame): Encoded DataFrame containing categorical features.
        seq_len (int): Length of input sequence for prediction.

    Returns:
        Tuple of (input_sequence, target_labels)
    """
    def __init__(self, df, seq_len):
        self.seq_len = seq_len
        # Prepare sequences of length `seq_len` and their next-step targets
        self.data = [
            (df.iloc[i:i+seq_len][categorical_cols].values,  # input sequence
             df.iloc[i+seq_len][categorical_cols].values)    # target
            for i in range(len(df) - seq_len)
        ]

    def __len__(self):
        return len(self.data)

    def __getitem__(self, idx):
        x_seq, y_target = self.data[idx]
        x_seq = torch.tensor(x_seq, dtype=torch.long)
        y_target = torch.tensor(y_target.astype(int), dtype=torch.long)
        return x_seq, y_target

# Path to the Excel file for logging results
excel_path = "risultati_modelli.xlsx"

def write_results_tocsv(model, Dropout, epochs, n_righe, losses, metrics, valoreThreshold):
    """
    Writes a new row of model results into an Excel file.

    Args:
        model: Trained model object (used to get model name).
        Dropout (float): Dropout rate used in the model.
        epochs (int): Number of training epochs.
        n_righe (int): Number of training/testing rows used.
        losses (list): List of loss values during training.
        metrics (dict): Dictionary of evaluation metrics (TP, FP, TN, FN, accuracy, etc.).
        valoreThreshold (float): Threshold that yielded the best performance.

    Creates the file with headers if it doesn't exist.
    Appends the result row to the file.
    """
    nuova_riga = [
        str(model.__class__.__name__),
        Dropout,
        epochs,
        n_righe,                        # Number of rows used
        round(losses[-1], 4),           # Final loss value
        metrics["TP"],
        metrics["FP"],
        metrics["TN"],
        metrics["FN"],
        round(valoreThreshold, 2),
        round(metrics["accuracy"], 4),
        round(metrics["precision"], 4),
        round(metrics["recall"], 4),
        round(metrics["f1"], 4),
    ]

    try:
        # Try opening existing Excel workbook
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    except FileNotFoundError:
        # If it doesn't exist, create a new one with headers
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Modello", "Drop", "Epoch", "RigheTrainTest", "Loss",
            "TP", "FP", "TN", "FN", "BestThreshold",
            "Accuracy", "Precision", "Recall", "f1-score"
        ])

    # Append the new row and save the file
    sheet.append(nuova_riga)
    workbook.save(excel_path)

def plot_istogramma(scores):
    """
    Plots a histogram  of anomaly scores.

    Args:
        scores (list or array): Anomaly scores to visualize.
    """
    plt.figure(figsize=(10, 6))
    sns.histplot(scores, bins=30, kde=True, color='skyblue', edgecolor='black')
    plt.title('Distribuzione degli Anomaly Scores')
    plt.xlabel('Anomaly Score')
    plt.ylabel('Frequenza / Densità')
    plt.grid(True)
    plt.show()

def plot_roc(tpr_list, fpr_list):
    """
    Plots the Receiver Operating Characteristic (ROC) curve.

    Args:
        tpr_list (list): True Positive Rates.
        fpr_list (list): False Positive Rates.

    Computes AUC and visualizes ROC curve for performance evaluation.
    """
    # Sort FPR and TPR in ascending FPR order to plot correctly
    sorted_indices = sorted(range(len(fpr_list)), key=lambda i: fpr_list[i])
    fpr_sorted = [fpr_list[i] for i in sorted_indices]
    tpr_sorted = [tpr_list[i] for i in sorted_indices]

    # Calculate Area Under Curve (AUC)
    roc_auc = auc(fpr_sorted, tpr_sorted)

    plt.figure(figsize=(8, 6))
    plt.plot(fpr_sorted, tpr_sorted, label=f'ROC Curve (AUC = {roc_auc:.2f})', linewidth=2.5)
    plt.plot([0, 1], [0, 1], linestyle='--', color='gray')  # Diagonal baseline
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('ROC Curve')
    plt.legend(loc='lower right')
    plt.grid(True)
    plt.show()
